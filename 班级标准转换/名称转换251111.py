import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil

"""
脚本作用：遍历文档中的每一格，
将不标准的班级名改为标准全称。

DeepSeek R1制作。
"""

class ClassNameStandardizer:
    def __init__(self):
        # 专业全称和简称映射
        self.major_mapping = {
            # 标准全称: [简称列表]
            '木材科学与工程': ['木材科学与工程卓越', '木材科学与工程', '木工卓越', '木工', '木材', '卓越', '木卓'],
            '林业工程类“成栋班”': ['林工程类成栋', '林工程成栋', '林工程', '木工成栋', '成栋'],
            '材料类': ['材料类', '材料'],
            '高分子材料与工程': ['高分子材料与工程', '高分子工程', '高分子材料', '高分子'],
            '材料化学': ['材料化学', '材化'],
            '轻化工程': ['轻化工程', '轻化'],
            '林产化工': ['林产化工', '林化']
        }
        
        # 创建反向映射
        self.reverse_major_mapping = {}
        for full_name, aliases in self.major_mapping.items():
            for alias in aliases:
                self.reverse_major_mapping[alias] = full_name
        
        # 汉字数字到阿拉伯数字映射
        self.chinese_number_map = {
            '一': '1', '二': '2', '三': '3', '四': '4', '五': '5',
            '六': '6', '七': '7', '八': '8', '九': '9'
        }
        
        # 特殊班级名称映射
        self.special_class_map = {
            '卓越': '卓越',
            '木卓': '卓越',
            '成栋': '1',
            '林业工程': '1'
        }

    def clean_text(self, text):
        """清理文本，去除空格换行等"""
        if pd.isna(text):
            return ""
        return str(text).strip().replace('\n', '').replace('\r', '').replace(' ', '')

    def extract_grade(self, text):
        """提取年级信息"""
        # 匹配4位数字的年份
        year_match = re.search(r'20(\d{2})', text)
        if year_match:
            return year_match.group(1)  # 返回后两位
        
        # 匹配2位数字的年级
        grade_match = re.search(r'(\d{2})级', text)
        if grade_match:
            return grade_match.group(1)
        
        # 匹配单独的2位数字（不在末尾）
        standalone_match = re.search(r'(\d{2})(?![年班级\d]|$)', text)
        if standalone_match:
            return standalone_match.group(1)
        
        return 'XX'

    def extract_class_number(self, text):
        """提取班级信息"""
        # 先检查特殊班级
        for special_key, special_value in self.special_class_map.items():
            if special_key in text:
                return special_value
        
        # 匹配"班"前的阿拉伯数字（1-2位）
        class_before_ban = re.search(r'(\d{1,2})班', text)
        if class_before_ban:
            class_num = class_before_ban.group(1)
            return str(int(class_num))  # 去除前导0
        
        # 匹配"班"前的汉字数字
        for chinese_num, arabic_num in self.chinese_number_map.items():
            if f"{chinese_num}班" in text:
                return arabic_num
        
        # 匹配以0开头的2位阿拉伯数字
        zero_lead_match = re.search(r'0(\d)', text)
        if zero_lead_match:
            return zero_lead_match.group(1)
        
        # 匹配末尾的单个阿拉伯数字
        end_digit_match = re.search(r'(\d)$', text)
        if end_digit_match:
            return end_digit_match.group(1)
        
        # 匹配末尾的单个汉字数字
        for chinese_num, arabic_num in self.chinese_number_map.items():
            if text.endswith(chinese_num):
                return arabic_num
        
        return 'Y'

    def identify_major(self, text):
        """识别专业名称"""
        # 先去除数字和常见干扰字符
        clean_text = re.sub(r'[\d年月日级班专业一二三四五六七八九\-_“”\s]', '', text)
        
        # 按长度排序，优先匹配长的简称
        sorted_aliases = sorted(self.reverse_major_mapping.keys(), 
                              key=len, reverse=True)
        
        for alias in sorted_aliases:
            if alias == clean_text:
                return self.reverse_major_mapping[alias]
        
        return None

    def standardize_class_name(self, text):
        """标准化班级名称"""
        if not text:
            return text
        
        clean_text = self.clean_text(text)
        if not clean_text:
            return text
        
        # 识别专业
        major = self.identify_major(clean_text)
        if not major:
            return text  # 不是班级名称，返回原文本
        
        # 提取年级和班级
        grade = self.extract_grade(clean_text)
        class_num = self.extract_class_number(clean_text)
        
        # 构建标准格式
        standardized = f"{major}{grade}-{class_num}"
        
        return standardized

    def process_excel_file(self, input_path, output_path):
        """处理单个Excel文件"""
        try:
            # 读取Excel文件
            wb = load_workbook(input_path)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            original_value = str(cell.value)
                            standardized = self.standardize_class_name(original_value)
                            
                            # 只有在确实识别为班级名称时才替换
                            if standardized != original_value and self.identify_major(self.clean_text(original_value)):
                                cell.value = standardized
            
            # 保存文件
            wb.save(output_path)
            print(f"处理完成: {os.path.basename(input_path)}")
            
        except Exception as e:
            print(f"处理文件 {input_path} 时出错: {str(e)}")
            # 如果出错，直接复制原文件
            shutil.copy2(input_path, output_path)

    def process_all_files(self):
        """处理所有Excel文件"""
        input_folder = "输入"
        output_folder = "输出"
        
        # 创建文件夹
        os.makedirs(input_folder, exist_ok=True)
        os.makedirs(output_folder, exist_ok=True)
        
        # 检查输入文件夹是否存在
        if not os.path.exists(input_folder):
            print(f"输入文件夹 '{input_folder}' 不存在，已创建空文件夹")
            return
        
        # 处理所有xlsx文件
        processed_count = 0
        for filename in os.listdir(input_folder):
            if filename.endswith('.xlsx'):
                input_path = os.path.join(input_folder, filename)
                output_path = os.path.join(output_folder, filename)
                
                self.process_excel_file(input_path, output_path)
                processed_count += 1
        
        if processed_count == 0:
            print(f"在 '{input_folder}' 文件夹中没有找到.xlsx文件")
        else:
            print(f"处理完成！共处理 {processed_count} 个文件，结果保存在 '{output_folder}' 文件夹中")

def main():
    """主函数"""
    print("=" * 30)
    print("开始处理班级名称标准化...")
    print("=" * 30)
    
    standardizer = ClassNameStandardizer()
    
    # 测试一些例子
    test_cases = [
        "25材料类一班",
        "24材料4班",
        "2024材料类四班",
        "25级轻化工程二班",
        "2025轻化2",
        "25木材科学与工程专业三班",
        "25木工卓越",
        "25木工卓越班",
        "23林产化工一班",
        "23林化一",
        "材料四班",
        "材料类4班",
        "材料类四班",
        "林产化工二班",
        "23林业工程“成栋班”",
        "25木卓",
        "大一材料类二班"  # 这个会比较难处理
    ]
    
    print("测试样例:")
    print("-" * 30)
    for test_case in test_cases:
        result = standardizer.standardize_class_name(test_case)
        print(f"输入: {test_case}")
        print(f"输出: {result}")
        print("-" * 30)
    
    print("\n开始处理Excel文件...")
    standardizer.process_all_files()

if __name__ == "__main__":
    main()
    input("\n按回车键退出...\n")
    